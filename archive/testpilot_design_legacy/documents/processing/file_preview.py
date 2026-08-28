#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
文件预览工具模块
支持多种文件格式的预览，包括PDF、DOCX、JSON、CSV、XLSX、YAML等
"""
import json
import csv
from pathlib import Path
from typing import Dict, Any

try:
    import pandas as pd
    PANDAS_SUPPORT = True
except ImportError:
    pd = None  # type: ignore
    PANDAS_SUPPORT = False

try:
    from docx import Document
    DOCX_SUPPORT = True
except ImportError:
    Document = None  # type: ignore
    DOCX_SUPPORT = False

try:
    from PyPDF2 import PdfReader
    PDF_SUPPORT = True
except ImportError:
    PdfReader = None  # type: ignore
    PDF_SUPPORT = False

try:
    import openpyxl
    XLSX_SUPPORT = True
except ImportError:
    openpyxl = None  # type: ignore
    XLSX_SUPPORT = False

try:
    import yaml
    YAML_SUPPORT = True
except ImportError:
    yaml = None  # type: ignore
    YAML_SUPPORT = False

from utils.utils_core.logger import get_logger

logger = get_logger(__name__)


def get_file_preview(file_path: str, max_length: int = 3000, max_csv_rows: int = 50) -> Dict[str, Any]:  # type: ignore[return-value]
    """
    获取文件预览内容
    
    Args:
        file_path: 文件路径
        max_length: 文本预览最大长度（字符数）
        max_csv_rows: CSV/Excel预览最大行数
    
    Returns:
        Dict包含:
            - content: 预览内容（文本或HTML）
            - content_type: 'text' 或 'html' 或 'json'
            - file_type: 文件扩展名
            - is_truncated: 是否被截断
            - error: 错误信息（如果有）
    """
    file_path_obj = Path(file_path)
    file_type = 'unknown'  # 初始化变量，避免可能的未定义错误
    
    if not file_path_obj.exists():
        return {
            'content': f'文件不存在: {file_path_obj.name}',
            'content_type': 'text',
            'file_type': file_path_obj.suffix[1:] if file_path_obj.suffix else 'unknown',
            'is_truncated': False,
            'error': 'File not found'
        }
    
    ext = file_path_obj.suffix.lower()
    file_type = ext[1:] if ext else 'unknown'
    
    try:
        # ----------------- 文本 / Markdown -----------------
        if ext in ['.txt', '.md']:
            with open(file_path_obj, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            is_truncated = len(content) > max_length
            preview = content[:max_length] + ('...' if is_truncated else '')
            
            return {
                'content': preview,
                'content_type': 'text',
                'file_type': file_type,
                'is_truncated': is_truncated
            }
        
        # ----------------- JSON -----------------
        elif ext == '.json':
            with open(file_path_obj, 'r', encoding='utf-8', errors='ignore') as f:
                try:
                    data = json.load(f)
                    # 格式化为易读的JSON字符串
                    formatted_json = json.dumps(data, ensure_ascii=False, indent=2)
                    
                    is_truncated = len(formatted_json) > max_length
                    preview = formatted_json[:max_length] + ('...' if is_truncated else '')
                    
                    return {
                        'content': preview,
                        'content_type': 'json',  # 前端可以用JSON语法高亮
                        'file_type': file_type,
                        'is_truncated': is_truncated
                    }
                except json.JSONDecodeError as e:
                    # JSON解析失败，尝试作为文本读取
                    f.seek(0)
                    content = f.read()
                    is_truncated = len(content) > max_length
                    preview = content[:max_length] + ('...' if is_truncated else '')
                    return {
                        'content': f'[JSON解析失败，显示原始内容]\n{preview}',
                        'content_type': 'text',
                        'file_type': file_type,
                        'is_truncated': is_truncated,
                        'error': f'JSON解析错误: {str(e)}'
                    }
        
        # ----------------- CSV -----------------
        elif ext == '.csv':
            if not PANDAS_SUPPORT:
                    # 如果没有pandas，使用标准csv库
                with open(file_path_obj, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    rows = []
                    for i, row in enumerate(reader):
                        if i >= max_csv_rows:
                            break
                        rows.append(row)
                    
                    # 转换为HTML表格
                    if rows:
                        html_rows = []
                        row_count = 0
                        for i, row in enumerate(rows):
                            row_count = i
                            if i == 0:
                                cells = ''.join([f'<th>{_escape_html(str(cell))}</th>' for cell in row])
                            else:
                                cells = ''.join([f'<td>{_escape_html(str(cell))}</td>' for cell in row])
                            html_rows.append(f'<tr>{cells}</tr>')
                        
                        is_truncated = row_count >= max_csv_rows - 1
                        html_content = f'<table class="table table-sm table-bordered"><tbody>{"".join(html_rows)}</tbody></table>'
                        if is_truncated:
                            html_content += f'<p class="text-muted small mt-2">... (仅显示前 {max_csv_rows} 行)</p>'
                        
                        return {
                            'content': html_content,
                            'content_type': 'html',
                            'file_type': file_type,
                            'is_truncated': is_truncated
                        }
                    else:
                        return {
                            'content': 'CSV文件为空',
                            'content_type': 'text',
                            'file_type': file_type,
                            'is_truncated': False
                        }
            else:
                # 使用pandas读取
                try:
                    if pd is None:
                        raise ImportError("pandas not installed")
                    df = pd.read_csv(str(file_path), nrows=max_csv_rows)
                    is_truncated = len(df) >= max_csv_rows
                    
                    # 转换为HTML表格
                    html_content = df.to_html(
                        classes='table table-sm table-bordered',
                        table_id='csv-preview-table',
                        escape=False,
                        index=False
                    )
                    if is_truncated:
                        html_content += f'<p class="text-muted small mt-2">... (仅显示前 {max_csv_rows} 行)</p>'
                    
                    return {
                        'content': html_content,
                        'content_type': 'html',
                        'file_type': file_type,
                        'is_truncated': is_truncated
                    }
                except Exception as e:
                    logger.warning(f"使用pandas读取CSV失败，尝试标准库: {e}")
                    # 回退到标准csv库
                    with open(file_path_obj, 'r', encoding='utf-8', errors='ignore') as f:
                        reader = csv.reader(f)
                        rows = [row for i, row in enumerate(reader) if i < max_csv_rows]
                        if rows:
                            html_rows = []
                            for i, row in enumerate(rows):
                                if i == 0:
                                    cells = ''.join([f'<th>{_escape_html(str(cell))}</th>' for cell in row])
                                else:
                                    cells = ''.join([f'<td>{_escape_html(str(cell))}</td>' for cell in row])
                                html_rows.append(f'<tr>{cells}</tr>')
                            
                            is_truncated = len(rows) >= max_csv_rows
                            html_content = f'<table class="table table-sm table-bordered"><tbody>{"".join(html_rows)}</tbody></table>'
                            if is_truncated:
                                html_content += f'<p class="text-muted small mt-2">... (仅显示前 {max_csv_rows} 行)</p>'
                            
                            return {
                                'content': html_content,
                                'content_type': 'html',
                                'file_type': file_type,
                                'is_truncated': is_truncated
                            }
        
        # ----------------- Word (DOCX) -----------------
        elif ext == '.docx':
            if not DOCX_SUPPORT:
                return {
                    'content': 'DOCX解析库未安装，无法预览此文件',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': 'python-docx not installed'
                }
            
            try:
                if Document is None:
                    raise ImportError("python-docx not installed")
                doc = Document(str(file_path_obj))
                paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
                content = '\n\n'.join(paragraphs)
                
                is_truncated = len(content) > max_length
                preview = content[:max_length] + ('...' if is_truncated else '')
                
                return {
                    'content': preview,
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': is_truncated
                }
            except Exception as e:
                return {
                    'content': f'DOCX解析失败: {str(e)}',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': str(e)
                }
        
        # ----------------- PDF -----------------
        elif ext == '.pdf':
            if not PDF_SUPPORT:
                return {
                    'content': 'PDF解析库未安装，无法预览此文件',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': 'PyPDF2 not installed'
                }
            
            try:
                if PdfReader is None:
                    raise ImportError("PyPDF2 not installed")
                reader = PdfReader(str(file_path_obj))
                pages_text = []
                total_chars = 0
                page_num = 0
                total_pages = len(reader.pages)
                
                for page_num, page in enumerate(reader.pages, start=1):
                    if total_chars >= max_length:
                        break
                    text = page.extract_text() or ""
                    if text.strip():
                        pages_text.append(f'[第 {page_num} 页]\n{text}')
                        total_chars += len(text)
                
                content = '\n\n'.join(pages_text)
                is_truncated = total_chars >= max_length or page_num < total_pages
                preview = content[:max_length] + ('...' if is_truncated else '')
                
                return {
                    'content': preview,
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': is_truncated
                }
            except Exception as e:
                return {
                    'content': f'PDF解析失败: {str(e)}',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': str(e)
                }
        
        # ----------------- Excel (XLSX) -----------------
        elif ext == '.xlsx':
            if not XLSX_SUPPORT:
                return {
                    'content': 'Excel解析库未安装，无法预览此文件',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': 'openpyxl not installed'
                }
            
            try:
                if openpyxl is None:
                    raise ImportError("openpyxl not installed")
                wb = openpyxl.load_workbook(str(file_path_obj), read_only=True)
                html_tables = []
                total_rows = 0
                
                for sheet_name in wb.sheetnames[:3]:  # 最多显示3个工作表
                    if total_rows >= max_csv_rows:
                        break
                    ws = wb[sheet_name]
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if i > max_csv_rows:
                            break
                        rows.append([str(cell) if cell is not None else '' for cell in row])
                        total_rows += 1
                    
                    if rows:
                        html_rows = []
                        for i, row in enumerate(rows):
                            if i == 0:
                                cells = ''.join([f'<th>{_escape_html(cell)}</th>' for cell in row])
                            else:
                                cells = ''.join([f'<td>{_escape_html(cell)}</td>' for cell in row])
                            html_rows.append(f'<tr>{cells}</tr>')
                        
                        html_table = f'<h6 class="mt-3">{_escape_html(sheet_name)}</h6><table class="table table-sm table-bordered"><tbody>{"".join(html_rows)}</tbody></table>'
                        html_tables.append(html_table)
                
                is_truncated = total_rows >= max_csv_rows
                html_content = ''.join(html_tables)
                if is_truncated:
                    html_content += f'<p class="text-muted small mt-2">... (仅显示前 {max_csv_rows} 行)</p>'
                
                return {
                    'content': html_content,
                    'content_type': 'html',
                    'file_type': file_type,
                    'is_truncated': is_truncated
                }
            except Exception as e:
                return {
                    'content': f'Excel解析失败: {str(e)}',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': str(e)
                }
        
        # ----------------- YAML -----------------
        elif ext in ['.yaml', '.yml']:
            if not YAML_SUPPORT:
                return {
                    'content': 'YAML解析库未安装，无法预览此文件',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': 'pyyaml not installed'
                }
            
            try:
                if yaml is None:
                    raise ImportError("pyyaml not installed")
                with open(file_path_obj, 'r', encoding='utf-8', errors='ignore') as f:
                    data = yaml.safe_load(f)
                    if data is None:
                        formatted_yaml = ''
                    else:
                        formatted_yaml = yaml.dump(data, allow_unicode=True, default_flow_style=False)
                    
                    is_truncated = len(formatted_yaml) > max_length
                    preview = formatted_yaml[:max_length] + ('...' if is_truncated else '')
                    
                    return {
                        'content': preview,
                        'content_type': 'text',
                        'file_type': file_type,
                        'is_truncated': is_truncated
                    }
            except Exception as e:
                return {
                    'content': f'YAML解析失败: {str(e)}',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': str(e)
                }
        
        # ----------------- 其他格式，尝试文本读取 -----------------
        else:
            try:
                with open(file_path_obj, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                
                is_truncated = len(content) > max_length
                preview = content[:max_length] + ('...' if is_truncated else '')
                
                return {
                    'content': preview,
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': is_truncated
                }
            except Exception as e:
                return {
                    'content': f'无法预览此文件格式 ({file_type})，错误: {str(e)}',
                    'content_type': 'text',
                    'file_type': file_type,
                    'is_truncated': False,
                    'error': str(e)
                }
    
    except Exception as e:
        logger.error(f"获取文件预览失败 {file_path_obj}: {e}", exc_info=True)
        # 确保 file_type 在异常情况下也有值
        if 'file_type' not in locals():
            file_type = 'unknown'
        return {
            'content': f'获取预览失败: {str(e)}',
            'content_type': 'text',
            'file_type': file_type,
            'is_truncated': False,
            'error': str(e)
        }


def _escape_html(text: str) -> str:
    """转义HTML特殊字符"""
    if not isinstance(text, str):
        text = str(text)
    return (text
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;')
            .replace("'", '&#x27;'))

