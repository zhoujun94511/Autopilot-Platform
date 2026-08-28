"""执行引擎核心。

遍历 TestCase 的 shell（before→case→after→fault），逐步骤派发关键字：
  - 解析每个 param 的实际值（变量/列绑定/对象库）
  - 按 keyword_id 在 REGISTRY 找实现并调用
  - 关键字若声明 out_params，把返回值写回变量池
  - 条件步骤 / 步骤组递归执行
  - 未实现关键字（被砍的 SAP/RSF 等）抛 NotImplementedKeyword，按策略记错继续

支持：顺序执行 + if/if_else 条件（param1/param2/expResult 算子）+ stepset 展开
+ 循环块（keyword_loop_start/end、mobile_loop_*：定次 cycle_times 或数据池 Excel/CSV/JSONArray 按行）
+ COLUMN 数据列绑定 + DataConfig 基线变量。
"""

from __future__ import annotations

import base64
import csv
import json
import logging
import os
import re
import threading
import time
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Optional

from ..model.testcase import (
    Shell,
    Step,
    StepSet,
    StepVerbs,
    StepInnerCase,
    TestCase,
)
from ..model import serializer
from ..model.loader import load_testcase
from ..keywords.registry import REGISTRY, NotImplementedKeyword
from ..keywords.context import ExecutionContext
from ..report.fail_class import classify_attribution, classify_failure
from ..runtime.log import get_logger
from .keyword_store import KeywordStore, discover_keywords
from .interrupt import RunInterrupted, bind_run_control
from .run_control import checkpoint
from .teardown import iter_teardown_steps


# 循环标记关键字 id（扁平：start..end 之间为循环体）
LOOP_START_IDS = {"keyword_loop_start", "mobile_loop_start"}
LOOP_END_IDS = {"keyword_loop_end", "mobile_loop_end"}
# else 标记（条件步骤子步骤中分隔 if/else 体）
ELSE_IDS = {"else"}


class FaultStrategy(Enum):
    STOP = "stop"          # 出错即停
    CONTINUE = "continue"  # 出错记录后继续


class StepTimeout(Exception):
    """单步执行超过引擎级最大时长（熔断，防单步 hang 死整个 run）。"""


# 引擎级单步最大执行时长(ms)。默认 0=不限（关闭）：超时靠守护线程实现，而 DB 连接/
# WebDriver 等有线程亲和性的资源在子线程里会失效，最可能 hang 的步骤又恰是这类——
# 故不能默认开启，仅作「明确知道代价」时的可选安全网（对齐参考实现 MaxExecuteTime=180000）。
DEFAULT_STEP_TIMEOUT_MS = 0


@dataclass
class StepResult:
    keyword_id: str
    comment: str          # 关键字通用说明（报告「说明」列）
    status: str          # PASS / FAIL / SKIP / NOIMPL / CANCEL
    message: str = ""
    duration_ms: int = 0
    loop_index: Optional[int] = None   # 数据驱动/循环里的迭代号（从 1 起；非循环内为 None）
    screenshot: str = ""               # 失败时抓取的截图(base64 PNG)，供报告内嵌
    remark: str = ""                   # 用例步骤备注（报告「备注」列）
    intent_id: str = ""
    binding_hit: str = ""              # cache | resolved | healed | failed | rolled_back
    heal_applied: bool = False
    resolved_keyword_id: str = ""      # intent 解析到的真实关键字（非步骤壳 keyword_id）
    fail_reason: str = ""              # no_candidate | element_not_found | timeout | ...
    fail_reason_label: str = ""
    rolled_back: bool = False
    resolve_strategy: str = ""         # cache | heuristic | heal | vision | ...
    candidate_count: int = 0
    perception_platform: str = ""
    perception_element_count: int = 0
    perception_used_screenshot: bool = False
    latency_ms: int = 0                # intent 解析+执行耗时（非整个步骤墙钟时可覆盖）
    vision_tokens: int = 0
    verification_status: str = ""      # passed | failed | skipped | missing
    screenshot_path: str = ""          # D3：失败后截图路径（after，不得写入 before）
    screenshot_before: str = ""        # 动作前截图 base64；禁止用 after 回填
    screenshot_before_path: str = ""   # 动作前截图路径
    dom_path: str = ""                 # D3：可选 page source / 控件树路径
    http_status: int = 0               # HTTP 步骤：响应码（有则写入报告）
    http_url: str = ""                 # HTTP 步骤：最终 URL
    http_elapsed_ms: int = 0           # HTTP 步骤：请求耗时
    fail_class: str = ""               # assertion | timeout | environment | locator | other
    fail_class_label: str = ""
    attribution: str = ""              # product_bug | env_issue | inner_agent_bug | tooling_gap | uncertain
    attribution_label: str = ""


@dataclass
class RunResult:
    case_name: str
    results: list[StepResult] = field(default_factory=list)
    source_path: str = ""          # 用例文件路径（报告展示/定位）
    platform: str = ""             # android / ios / 空=未标
    tag: str = ""                  # 用例标签（WEB/MOBILE 等）
    duration_ms: int = 0           # 用例总耗时
    device_udid: str = ""          # 并行执行时绑定的设备 UDID
    worker_slot: int = -1          # 并行 worker 槽位（-1=串行/未标）

    @property
    def passed(self) -> bool:
        # 仅 FAIL 判为用例失败；NOIMPL（有意砍除的关键字）/SKIP 不算失败（报告中仍单独可见）
        return not any(r.status == "FAIL" for r in self.results)

    def counts(self) -> dict[str, int]:
        c: dict[str, int] = {}
        for r in self.results:
            c[r.status] = c.get(r.status, 0) + 1
        return c


def _step_remark(node) -> str:
    return str(getattr(node, "remark", "") or "")


class Executor:
    def __init__(
        self,
        context: Optional[ExecutionContext] = None,
        fault_strategy: FaultStrategy = FaultStrategy.CONTINUE,
        keyword_store: Optional[KeywordStore] = None,
        cancel_event=None,
        pause_event=None,
        on_step=None,
        step_timeout_ms: int = DEFAULT_STEP_TIMEOUT_MS,
    ) -> None:
        self.ctx = context or ExecutionContext()
        self.fault_strategy = fault_strategy
        self.keyword_store = keyword_store
        self.step_timeout_ms = step_timeout_ms   # 单步执行超时熔断(ms)，0=不限
        self.cancel_event = cancel_event   # threading.Event；置位则在步骤间协作式停止
        self.pause_event = pause_event     # threading.Event；置位则在步骤间协作式暂停
        self.on_step = on_step             # 回调(StepResult)，用于异步进度
        self._ks_stack: set[str] = set()        # 自定义关键字递归防护
        self._innercase_stack: set[str] = set()  # 内嵌用例递归防护
        self._case_dir_stack: list[str] = []     # 当前用例所在目录（解析内嵌相对路径）
        self._loop_stack: list[int] = []         # 当前数据驱动/循环迭代号栈（报告 loop_index）
        # 当前步动作前截图；(b64, path)。FAIL 后另采 after，禁止用 after 覆盖。
        self._pre_action_shot: tuple[str, str] = ("", "")

    def _consume_intent_meta(self) -> dict:
        meta = self.ctx.get_var("__last_intent_meta__")
        self.ctx.set_var("__last_intent_meta__", None)
        if not isinstance(meta, dict):
            return {}
        def _int(key: str) -> int:
            try:
                return max(0, int(meta.get(key) or 0))
            except (TypeError, ValueError):
                return 0

        return {
            "intent_id": str(meta.get("intent_id") or ""),
            "binding_hit": str(meta.get("binding_hit") or ""),
            "heal_applied": bool(meta.get("heal_applied")),
            "resolved_keyword_id": str(meta.get("keyword_id") or ""),
            "fail_reason": str(meta.get("fail_reason") or ""),
            "fail_reason_label": str(meta.get("fail_reason_label") or ""),
            "rolled_back": bool(meta.get("rolled_back")),
            "resolve_strategy": str(meta.get("resolve_strategy") or ""),
            "candidate_count": _int("candidate_count"),
            "perception_platform": str(meta.get("perception_platform") or ""),
            "perception_element_count": _int("perception_element_count"),
            "perception_used_screenshot": bool(meta.get("perception_used_screenshot")),
            "latency_ms": _int("latency_ms"),
            "vision_tokens": _int("vision_tokens"),
            "verification_status": str(meta.get("verification_status") or ""),
            "screenshot_path": str(meta.get("screenshot_path") or ""),
            "dom_path": str(meta.get("dom_path") or ""),
        }

    # ---- 公开入口 ----
    def run_testcase(self, tc: TestCase) -> RunResult:
        # 独立运行时若未注入自定义关键字仓库，则就近从用例所在目录发现 .ks
        base_dir = (os.path.dirname(os.path.abspath(tc.source_path))
                    if tc.source_path else os.getcwd())
        if self.keyword_store is None and tc.source_path:
            self.keyword_store = discover_keywords(base_dir)
        self._case_dir_stack = [base_dir]
        # Intent Binding：注入逻辑用例 id 与运行平台
        lid = str(getattr(tc, "logical_case_id", "") or "").strip()
        if lid:
            self.ctx.set_var("__logical_case_id__", lid)
        plat = str(getattr(tc, "platform", "") or "").strip().lower()
        if plat in ("web", "android", "ios"):
            self.ctx.set_var("__run_platform__", plat)
        try:
            from ..intent.vision import reset_vision_call_budget  # 延迟：可选 Vision 配额

            reset_vision_call_budget()
        except (ImportError, AttributeError, RuntimeError):
            pass
        slot = self.ctx.get_var("__worker_slot__")
        try:
            worker_slot = int(slot) if slot not in (None, "") else -1
        except (TypeError, ValueError):
            worker_slot = -1
        result = RunResult(
            case_name=tc.name,
            source_path=tc.source_path or "",
            platform=(tc.platform or "").strip(),
            tag=(tc.tag or "").strip(),
            device_udid=str(self.ctx.get_var("__device_udid__") or ""),
            worker_slot=worker_slot,
        )
        start = time.time()
        bind_run_control(self.ctx, self.cancel_event, self.pause_event)
        # 壳语义（对齐参考实现四壳）：
        #   before 一次 → case（绑定数据池则整体逐行循环，否则一次）；
        #     STOP 策略下 before 失败则跳过 case。
        #   after 为 finally：无论成败都执行清理；用户取消时改跑精简版 after（白名单关键字）。
        #   fault 为失败兜底：仅当前面壳出现过失败时才执行。
        if not self._checkpoint():
            self._run_shell(tc.before, result)
        stop = not result.passed and self.fault_strategy is FaultStrategy.STOP
        if not stop and not self._checkpoint():
            self._run_case_shell(tc, result)
        if self._cancelled():
            self._run_teardown_shell(tc.after, result)
        elif not self._checkpoint():
            self._run_shell(tc.after, result)          # finally：总执行清理
        if not result.passed and not self._checkpoint():
            self._run_shell(tc.fault, result)          # 仅失败时兜底
        result.duration_ms = int((time.time() - start) * 1000)
        return result

    def _run_case_shell(self, tc: TestCase, result: RunResult) -> None:
        """case 主体：用例级绑定数据池 `DATATABLE(...)` 则整体逐行循环，否则执行一次。"""
        datapool = getattr(tc, "datapool", "") or ""
        try:
            rows = self._datatable_rows(datapool)
        except Exception as e:  # noqa: BLE001  绑定了但读不了：记 FAIL，仍跑一次主体
            self._add(result, StepResult(
                "用例数据池", tc.name, "FAIL", f"数据源读取失败: {e}"))
            rows = None
        if rows is None:
            self._run_shell(tc.case, result)
            return
        saved = dict(self.ctx.data_row)
        try:
            for idx, row in enumerate(rows):
                if self._checkpoint():
                    break
                self.ctx.data_row = {**saved, **row}
                self._loop_stack.append(idx + 1)
                try:
                    self._run_shell(tc.case, result)
                finally:
                    self._loop_stack.pop()
                if not result.passed and self.fault_strategy is FaultStrategy.STOP:
                    break
        finally:
            self.ctx.data_row = saved

    def _cancelled(self) -> bool:
        return self.cancel_event is not None and self.cancel_event.is_set()

    def _checkpoint(self) -> bool:
        return checkpoint(self.cancel_event, self.pause_event)

    # ---- 内部 ----
    def _run_shell(self, shell: Shell, result: RunResult) -> None:
        self._run_sequence(_group_loops(shell.steps), result)

    def _run_teardown_shell(self, shell: Shell, result: RunResult) -> None:
        steps = iter_teardown_steps(shell.steps)
        if not steps:
            return
        self._add(result, StepResult(
            "_teardown", "停止后精简清理", "PASS",
            f"执行 after 白名单 {len(steps)} 步"))
        for step in steps:
            if self._pause_hold():
                return
            self._run_step(step, result)

    def _pause_hold(self) -> bool:
        """精简清理阶段：不因已置位的 cancel 而跳过，仅响应暂停。"""
        return checkpoint(None, self.pause_event)

    def _run_sequence(self, items: list, result: RunResult) -> None:
        for item in items:
            if self._checkpoint():
                return
            if isinstance(item, LoopBlock):
                self._run_loop(item, result)
            else:
                self._run_node(item, result)
            if (
                self.fault_strategy is FaultStrategy.STOP
                and result.results
                and result.results[-1].status == "FAIL"
            ):
                return

    def _run_loop(self, block: "LoopBlock", result: RunResult) -> None:
        start = block.start
        body_items = _group_loops(block.body)
        loop_fail = (start.param("loop_failure_strategy")
                     or start.param("failStrategy") or "break")
        result.results.append(
            StepResult(start.keyword_id, start.comment, "PASS", "循环开始",
                       remark=_step_remark(start))
        )
        rows = self._iter_datapool(start)
        saved_row = dict(self.ctx.data_row)
        try:
            for idx, row in enumerate(rows):
                if self._checkpoint():
                    break
                if row is not None:
                    self.ctx.data_row = {**saved_row, **row}
                before = len(result.results)
                self._loop_stack.append(idx + 1)
                try:
                    self._run_sequence(body_items, result)
                finally:
                    self._loop_stack.pop()
                failed = any(r.status == "FAIL" for r in result.results[before:])
                if failed and loop_fail == "break":
                    result.results.append(
                        StepResult("loop", block.start.comment, "SKIP",
                                   f"第{idx + 1}轮失败，按 break 策略跳出循环",
                                   remark=_step_remark(block.start)))
                    break
        finally:
            self.ctx.data_row = saved_row
        result.results.append(StepResult("keyword_loop_end", "", "PASS", "循环结束"))

    def _iter_datapool(self, start: Step):
        """根据 loop_start 参数产出迭代数据行；定次循环产出 None 占位。"""
        src_type = (start.param("datapool_source_type") or "无").strip()
        path = self.ctx.resolve(start.param("datapool_source_path") or "") or ""
        if src_type in ("Excel", "CSV") and path and os.path.exists(str(path)):
            return _read_table(str(path))
        if src_type == "JSONArray" and path and os.path.exists(str(path)):
            with open(str(path), "r", encoding="utf-8") as f:
                data = json.load(f)
            return [dict(r) for r in data]
        # 定次循环
        n_raw = (self.ctx.resolve(start.param("cycle_times") or "")
                 or start.param("loopCount") or "1")
        try:
            n = int(float(str(n_raw)))
        except (TypeError, ValueError):
            n = 1
        return [None] * max(n, 0)

    def _run_node(self, node, result: RunResult) -> None:
        if isinstance(node, Step):
            if node.is_condition:
                self._run_condition(node, result)
            else:
                self._run_step(node, result)
        elif isinstance(node, StepSet):
            self._run_stepset(node, result)
        elif isinstance(node, StepVerbs):
            self._run_stepverbs(node, result)
        elif isinstance(node, StepInnerCase):
            self._run_innercase(node, result)

    def _datatable_rows(self, spec: str):
        """按数据池绑定 `DATATABLE(源,私有)` 读取数据行；未绑定/NONE 返回 None。

        源支持 .xlsx/.csv/.json；相对路径按 工程根 或 当前用例目录 解析；允许含变量 ${}。
        读取失败抛异常（由调用方决定记 FAIL 还是忽略）。
        """
        source, _private = _parse_datatable(spec)
        if not source or source.upper() == "NONE":
            return None
        path = str(self.ctx.resolve(source) or source)
        if not os.path.isabs(path):
            base = (self.ctx.get_var("__project_path__")
                    or (self._case_dir_stack[-1] if self._case_dir_stack else os.getcwd()))
            cand = os.path.join(str(base), path)
            if os.path.exists(cand):
                path = cand
        if not os.path.exists(path):
            raise FileNotFoundError(f"数据源文件未找到: {path}")
        if path.lower().endswith((".json",)):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return [dict(r) for r in data]
        return _read_table(path)

    def _run_stepset(self, node: StepSet, result: RunResult) -> None:
        """步骤组：绑定 datapool 则按数据行逐行执行 children（私有作用域，行末还原
        data_row）；未绑定/NONE 则执行一次。对齐参考实现 stepset 数据驱动。"""
        if not node.is_run:
            self._add(result, StepResult(
                f"组:{node.name}", node.comment, "SKIP", "isrun=false",
                remark=_step_remark(node)))
            return
        body = _group_loops(node.children)
        try:
            rows = self._datatable_rows(node.datapool)
        except Exception as e:  # noqa: BLE001  数据源读取失败记 FAIL，不崩溃整次执行
            self._add(result, StepResult(
                f"组:{node.name}", node.comment, "FAIL", f"数据源读取失败: {e}",
                remark=_step_remark(node)))
            return
        if rows is None:                       # 未绑定数据池：跑一次
            self._run_sequence(body, result)
            return
        saved = dict(self.ctx.data_row)
        try:
            for idx, row in enumerate(rows):
                if self._checkpoint():
                    break
                self.ctx.data_row = {**saved, **row}
                self._loop_stack.append(idx + 1)
                try:
                    self._run_sequence(body, result)
                finally:
                    self._loop_stack.pop()
                if not result.passed and self.fault_strategy is FaultStrategy.STOP:
                    break
        finally:
            self.ctx.data_row = saved          # 私有作用域：还原，不污染外层
        self._add(result, StepResult(
            f"组:{node.name}", node.comment, "PASS", f"数据驱动 {len(rows)} 行",
            remark=_step_remark(node)))

    def _run_stepverbs(self, node: StepVerbs, result: RunResult) -> None:
        """调用自定义关键字(.ks)：把其步骤序列内联展开执行。

        实参绑定为局部变量（执行后恢复，不污染外层作用域）；形参声明的默认值用于补缺；
        通过 _ks_stack 防止递归调用。
        """
        if not node.is_run:
            result.results.append(
                StepResult(node.ks_id, node.comment, "SKIP", "isrun=false",
                           remark=_step_remark(node)))
            return
        kdef = self.keyword_store.get(node.ks_id) if self.keyword_store else None
        if kdef is None:
            result.results.append(
                StepResult(node.ks_id, node.comment, "NOIMPL", "未找到自定义关键字定义(.ks)",
                           remark=_step_remark(node)))
            return
        if node.ks_id in self._ks_stack:
            result.results.append(
                StepResult(node.ks_id, node.comment, "SKIP", "检测到自定义关键字递归调用，跳过",
                           remark=_step_remark(node)))
            return

        # 调用处实参 → 局部变量；形参默认值补缺
        args = {p.param_id: self.ctx.resolve(p.value) for p in node.params}
        for lp in kdef.params:
            if lp.param_id not in args and lp.default:
                args[lp.param_id] = self.ctx.resolve(lp.default)
        scoped = list(args.keys())
        had = {n for n in scoped if n in self.ctx.variables}
        saved = {n: self.ctx.variables[n] for n in had}
        self.ctx.variables.update(args)

        result.results.append(
            StepResult(node.ks_id, node.comment, "PASS", f"调用自定义关键字 {kdef.ks_id}",
                       remark=_step_remark(node)))
        self._ks_stack.add(node.ks_id)
        # 不用 try/finally：执行体出错也要先还原作用域再抛出（语义同 finally，
        # 但避免静态分析对 try/finally 的可达性误判）
        error = None
        try:
            self._run_sequence(_group_loops(kdef.steps), result)
        except Exception as e:  # noqa: BLE001  捕获后还原作用域再原样抛出
            error = e
        self._ks_stack.discard(node.ks_id)
        for n in scoped:                       # 恢复局部作用域
            if n in had:
                self.ctx.variables[n] = saved[n]
            else:
                self.ctx.variables.pop(n, None)
        if error is not None:
            raise error

    def _run_innercase(self, node: StepInnerCase, result: RunResult) -> None:
        """内嵌引用另一个 .tc：加载后把其 before→case→after 步骤内联到当前执行流。

        relativepath 相对当前用例所在目录解析；通过 _innercase_stack 防止循环引用。
        共享当前 ExecutionContext（变量/对象库/driver）与自定义关键字仓库。
        """
        if not node.is_run:
            result.results.append(
                StepResult(node.relative_path, node.comment, "SKIP", "isrun=false",
                           remark=_step_remark(node)))
            return
        base = self._case_dir_stack[-1] if self._case_dir_stack else os.getcwd()
        path = os.path.normpath(os.path.join(base, node.relative_path))
        if not os.path.exists(path):
            result.results.append(
                StepResult(node.relative_path, node.comment, "FAIL",
                           f"内嵌用例文件未找到: {path}", remark=_step_remark(node)))
            return
        key = os.path.abspath(path)
        if key in self._innercase_stack:
            result.results.append(
                StepResult(node.relative_path, node.comment, "SKIP",
                           "检测到内嵌用例循环引用，跳过", remark=_step_remark(node)))
            return
        try:
            inner = _load_inner_case(path)
        # noinspection PyBroadException
        except Exception as e:  # noqa: BLE001  加载失败记 FAIL，不崩溃整次执行
            result.results.append(
                StepResult(node.relative_path, node.comment, "FAIL",
                           f"内嵌用例加载失败: {e}", remark=_step_remark(node)))
            return

        result.results.append(
            StepResult(node.relative_path, node.comment, "PASS", f"内嵌用例 {inner.name}",
                       remark=_step_remark(node)))
        self._innercase_stack.add(key)
        self._case_dir_stack.append(os.path.dirname(key))
        try:
            for shell in (inner.before, inner.case, inner.after):
                self._run_shell(shell, result)
        finally:
            self._case_dir_stack.pop()
            self._innercase_stack.discard(key)

    def _run_condition(self, step: Step, result: RunResult) -> None:
        """条件步骤：param1/param2 + expResult 算子求值；真→if 体，假→else 体。

        兼容早期样例的 condition 表达式参数（形如 a==b）。
        子步骤中的 `else` 标记分隔 if 体与 else 体。
        """
        if step.param("param1") is not None or step.param("param2") is not None:
            v1 = str(self.ctx.resolve(step.param("param1") or ""))
            v2 = str(self.ctx.resolve(step.param("param2") or ""))
            op = step.param("expResult") or "等于(精确匹配)"
            truthy = _compare(v1, v2, op)
            detail = f"[{v1}] {op} [{v2}] = {truthy}"
        else:
            cond = str(self.ctx.resolve(step.param("condition") or ""))
            truthy = _eval_expr(cond)
            detail = f"条件={truthy}"

        result.results.append(
            StepResult(step.keyword_id, step.comment, "PASS", detail,
                       remark=_step_remark(step))
        )
        if_body, else_body = _split_else(step.children)
        branch = if_body if truthy else else_body
        self._run_sequence(_group_loops(branch), result)

    def _consume_http_meta(self, keyword_id: str) -> dict[str, Any]:
        """http_* 步骤从 ctx.last_http 带走 status/url/耗时，避免污染后续步骤。"""
        kid = str(keyword_id or "")
        last = getattr(self.ctx, "last_http", None)
        if not kid.startswith("http_") or not isinstance(last, dict):
            return {}
        self.ctx.last_http = {}
        out: dict[str, Any] = {}
        try:
            status = int(last.get("status") or 0)
        except (TypeError, ValueError):
            status = 0
        if status:
            out["http_status"] = status
        url = str(last.get("url") or "").strip()
        if url:
            out["http_url"] = url[:500]
        try:
            elapsed = int(last.get("elapsed_ms") or 0)
        except (TypeError, ValueError):
            elapsed = 0
        if elapsed:
            out["http_elapsed_ms"] = elapsed
        return out

    def _enrich_step_result(self, sr: StepResult) -> None:
        http = self._consume_http_meta(sr.keyword_id)
        if http.get("http_status") and not sr.http_status:
            sr.http_status = int(http["http_status"])
        if http.get("http_url") and not sr.http_url:
            sr.http_url = str(http["http_url"])
        if http.get("http_elapsed_ms") and not sr.http_elapsed_ms:
            sr.http_elapsed_ms = int(http["http_elapsed_ms"])
        crash_pkg = ""
        if sr.status == "FAIL":
            from .app_watch import detect_crash_on_fail  # 延迟：失败才探前台，避免热路径顶栏

            crash_pkg = detect_crash_on_fail(self.ctx, sr.message)
            if crash_pkg:
                sr.fail_reason = "app_crash"
                sr.fail_reason_label = "应用崩溃"
                note = f"目标应用已离开前台（疑似崩溃，期望 {crash_pkg}）"
                if note not in (sr.message or ""):
                    sr.message = f"{note}。{sr.message}".strip("。")
                sr.attribution = "product_bug"
                sr.attribution_label = "产品缺陷"
        if sr.status == "FAIL" and (not sr.fail_class or crash_pkg):
            cls = classify_failure(
                keyword_id=sr.keyword_id,
                fail_reason=sr.fail_reason,
                message=sr.message,
                status=sr.status,
            )
            sr.fail_class = str(cls.get("fail_class") or "")
            sr.fail_class_label = str(cls.get("fail_class_label") or "")
        if sr.status == "FAIL" and not sr.attribution:
            attr = classify_attribution(
                fail_class=sr.fail_class,
                fail_reason=sr.fail_reason,
                resolve_strategy=sr.resolve_strategy,
                binding_hit=sr.binding_hit,
                heal_applied=sr.heal_applied,
                keyword_id=sr.keyword_id,
                message=sr.message,
                status=sr.status,
            )
            sr.attribution = str(attr.get("attribution") or "")
            sr.attribution_label = str(attr.get("attribution_label") or "")

    def _add(self, result: RunResult, sr: StepResult) -> None:
        """追加步骤结果并触发进度回调（供异步执行更新 UI）。"""
        self._enrich_step_result(sr)
        if sr.loop_index is None and self._loop_stack:
            sr.loop_index = self._loop_stack[-1]     # 标注当前(最内层)迭代号
        result.results.append(sr)
        # 落盘：每步结果进日志文件（GUI 已由 on_step→add_step 渲染，故 ap_no_gui 不重复进控制台）
        lvl = (logging.ERROR if sr.status == "FAIL"
               else logging.WARNING if sr.status in ("NOIMPL", "SKIP", "CANCEL")
               else logging.INFO)
        get_logger("engine").log(lvl, "[%s] %s | %s | %s", sr.status, sr.keyword_id,
                                 sr.comment or "", sr.message or "",
                                 extra={"ap_no_gui": True})
        if self.on_step is not None:
            self.on_step(sr)

    def _call_keyword(self, func, kwargs):
        """调用关键字实现；step_timeout_ms>0 时套一层超时熔断。

        超时无法强杀正在运行的关键字（Python 限制），故在守护线程里跑：超时即抛
        StepTimeout（由 _run_step 记为 FAIL 继续下一步），后台线程自行了结不阻塞 run。
        """
        timeout_ms = self.step_timeout_ms
        if not timeout_ms or timeout_ms <= 0:
            return func(self.ctx, **kwargs)
        box: dict = {}

        def _run():
            try:
                box["ret"] = func(self.ctx, **kwargs)
            except BaseException as e:  # noqa: BLE001  原样带回主线程
                box["err"] = e

        t = threading.Thread(target=_run, daemon=True)
        t.start()
        t.join(timeout_ms / 1000.0)
        if t.is_alive():
            raise StepTimeout(
                f"步骤执行超过 {timeout_ms}ms 未返回，已熔断"
                "（后台线程可能仍在运行）")
        if "err" in box:
            raise box["err"]
        return box.get("ret")

    def _run_step(self, step: Step, result: RunResult) -> None:
        if not step.is_run:
            self._add(result, StepResult(step.keyword_id, step.comment, "SKIP", "isrun=false",
                                         remark=_step_remark(step)))
            return

        kwdef = REGISTRY.get(step.keyword_id)
        if kwdef is None:
            self._add(result, StepResult(
                step.keyword_id, step.comment, "NOIMPL", "关键字未实现/已砍除",
                remark=_step_remark(step)))
            return

        start = time.monotonic()
        log_start = len(self.ctx.logs)
        self._stash_pre_action_shot(result, step)
        try:
            ret = self._execute_step_body(step, kwdef)
            # 输出变量回写
            if kwdef.out_params and isinstance(ret, dict):
                for name, val in ret.items():
                    self.ctx.set_var(name, val)
            elif kwdef.out_params and ret is not None:
                self.ctx.set_var(kwdef.out_params[0], ret)
            dur = int((time.monotonic() - start) * 1000)
            step_msg = "\n".join(self.ctx.logs[log_start:]).strip()
            meta = self._consume_intent_meta()
            self._add(result, StepResult(
                step.keyword_id, step.comment, "PASS", step_msg,
                duration_ms=dur, remark=_step_remark(step), **meta))
            self._pre_action_shot = ("", "")
        except NotImplementedKeyword as e:
            self._add(result, StepResult(step.keyword_id, step.comment, "NOIMPL", str(e),
                                         remark=_step_remark(step)))
        except RunInterrupted as e:
            dur = int((time.monotonic() - start) * 1000)
            self._add(result, StepResult(
                step.keyword_id, step.comment, "CANCEL", str(e), duration_ms=dur,
                remark=_step_remark(step)))
        # noinspection PyBroadException
        except Exception as e:  # noqa: BLE001  记录所有异常为失败（含 KeywordError）
            if self._try_ios_alert_retry(step, kwdef, start, log_start, result):
                return
            dur = int((time.monotonic() - start) * 1000)
            meta = self._consume_intent_meta()
            self._attach_fail_shots(meta, result, step)
            self._add(result, StepResult(
                step.keyword_id, step.comment, "FAIL", str(e), duration_ms=dur,
                screenshot=str(meta.pop("screenshot", "") or ""),
                screenshot_before=str(meta.pop("screenshot_before", "") or ""),
                remark=_step_remark(step), **meta))

    def _execute_step_body(self, step: Step, kwdef) -> object:
        kwargs = {}
        for p in step.params:
            kwargs[p.param_id] = self.ctx.resolve(p.value)
        return self._call_keyword(kwdef.func, kwargs)

    def _try_ios_alert_retry(self, step: Step, kwdef, start: float, log_start: int,
                             result: RunResult) -> bool:
        """步骤 FAIL 后：若存在 iOS 系统弹框则处理并重试一次。"""
        from ..runtime import settings as app_settings  # 延迟：仅 iOS 失败重试读设置
        from ..mobile.ios.alert import maybe_handle_ios_alert  # 延迟：可选 iOS 弹窗处理

        mgr = getattr(self.ctx, "appium", None)
        if mgr is None or getattr(mgr, "platform", "") != "ios":
            return False
        if not app_settings.ios_alert_enabled():
            return False
        ctx_val = self.ctx.get_var("__ios_alert_enabled__")
        if ctx_val is not None and not bool(ctx_val):
            return False
        if not app_settings.ios_alert_retry_on_handled():
            return False
        alert_res = maybe_handle_ios_alert(self.ctx, stage="on_error")
        if not alert_res.handled:
            return False
        self.ctx.log(
            f"iOS 系统弹框已处理({alert_res.action})，重试步骤: {step.keyword_id}"
        )
        try:
            ret = self._execute_step_body(step, kwdef)
            if kwdef.out_params and isinstance(ret, dict):
                for name, val in ret.items():
                    self.ctx.set_var(name, val)
            elif kwdef.out_params and ret is not None:
                self.ctx.set_var(kwdef.out_params[0], ret)
            dur = int((time.monotonic() - start) * 1000)
            step_msg = "\n".join(self.ctx.logs[log_start:]).strip()
            # Alert 重试成功也要带上 intent meta（result_json / 自动化回写依赖）
            meta = self._consume_intent_meta()
            self._add(result, StepResult(step.keyword_id, step.comment, "PASS", step_msg,
                                         duration_ms=dur, remark=_step_remark(step), **meta))
            return True
        # noinspection PyBroadException
        except Exception as retry_err:  # noqa: BLE001
            dur = int((time.monotonic() - start) * 1000)
            meta = self._consume_intent_meta()
            self._attach_fail_shots(meta, result, step)
            self._add(result, StepResult(
                step.keyword_id, step.comment, "FAIL", str(retry_err), duration_ms=dur,
                screenshot=str(meta.pop("screenshot", "") or ""),
                screenshot_before=str(meta.pop("screenshot_before", "") or ""),
                remark=_step_remark(step), **meta))
            return True

    @staticmethod
    def _needs_pre_screenshot(keyword_id: str) -> bool:
        kid = (keyword_id or "").lower()
        return any(
            token in kid
            for token in (
                "element",
                "click",
                "swipe",
                "type",
                "input",
                "verify",
                "assert",
                "intent",
                "locate",
                "tap",
            )
        )

    def _stash_pre_action_shot(self, result: RunResult, step: Any) -> None:
        """动作前采一帧；只写 before，后续 after 不得覆盖。"""
        self._pre_action_shot = ("", "")
        if not self._needs_pre_screenshot(getattr(step, "keyword_id", "") or ""):
            return
        b64, path, _dom = self._capture_evidence(
            result, step=step, intent_id="", filename="screenshot_before.png"
        )
        self._pre_action_shot = (b64, path)

    def _attach_fail_shots(self, meta: dict, result: RunResult, step: Any) -> None:
        """FAIL 后另采 after；before 只用动作前缓存，禁止用 after 回填。"""
        pre_b64, pre_path = self._pre_action_shot
        post_b64, post_path, dom_path = self._capture_evidence(
            result,
            step=step,
            intent_id=str(meta.get("intent_id") or ""),
            filename="screenshot.png",
        )
        if post_path:
            meta["screenshot_path"] = post_path
        if dom_path:
            meta["dom_path"] = dom_path
        if pre_path:
            meta["screenshot_before_path"] = pre_path
        meta["screenshot"] = post_b64
        meta["screenshot_before"] = pre_b64
        self._pre_action_shot = ("", "")

    def _capture_screenshot(self) -> str:
        """尽力抓当前设备/浏览器截图 → base64 PNG（供报告内嵌）；无 driver/失败返回空串。"""
        b64, _path, _dom = self._capture_evidence(None)
        return b64

    def _capture_evidence(
        self,
        result: RunResult | None,
        *,
        step: Any = None,
        intent_id: str = "",
        filename: str = "screenshot.png",
    ) -> tuple[str, str, str]:
        """抓截图（base64）并尽量落盘证据路径（D3）。

        返回 (base64_png, screenshot_path, dom_path)；路径相对工程根（若可知）。
        """
        ctx = self.ctx
        png: bytes | None = None
        driver = None

        def _mobile():
            mgr = getattr(ctx, "appium", None)
            return mgr.driver() if mgr is not None else None

        def _web():
            w = getattr(ctx, "web", None)
            return w.driver() if w is not None else None

        for getter in (_mobile, _web):
            try:
                drv = getter()
                if drv is not None and hasattr(drv, "get_screenshot_as_png"):
                    raw = drv.get_screenshot_as_png()
                    if raw:
                        png = raw
                        driver = drv
                        break
            except (OSError, RuntimeError, AttributeError, TypeError, ValueError):
                continue

        if not png:
            return "", "", ""

        b64 = base64.b64encode(png).decode("ascii")
        project = str(ctx.get_var("__project_path__") or "").strip()
        if not project:
            return b64, "", ""

        case_name = ""
        if result is not None:
            case_name = str(getattr(result, "case_name", "") or "case")
        safe_case = "".join(c if c.isalnum() or c in "-_." else "_" for c in case_name)[:80] or "case"
        sid = (intent_id or "").strip()
        if not sid and step is not None:
            sid = str(getattr(step, "keyword_id", "") or "step")
        safe_sid = "".join(c if c.isalnum() or c in "-_." else "_" for c in sid)[:64] or "step"
        stamp = time.strftime("%Y%m%dT%H%M%S")
        ev_dir = Path(project) / "reports" / "evidence" / safe_case / f"{safe_sid}_{stamp}"
        try:
            ev_dir.mkdir(parents=True, exist_ok=True)
            safe_name = (filename or "screenshot.png").replace("\\", "/").split("/")[-1]
            if not safe_name.lower().endswith((".png", ".jpg", ".jpeg")):
                safe_name = "screenshot.png"
            shot_file = ev_dir / safe_name
            shot_file.write_bytes(png)
            try:
                shot_rel = str(shot_file.relative_to(Path(project))).replace("\\", "/")
            except ValueError:
                shot_rel = str(shot_file)
            dom_rel = ""
            # 可选 DOM / page source
            if driver is not None and hasattr(driver, "page_source"):
                try:
                    src = driver.page_source
                    if src:
                        dom_file = ev_dir / "page_source.xml"
                        dom_file.write_text(str(src)[:2_000_000], encoding="utf-8")
                        try:
                            dom_rel = str(dom_file.relative_to(Path(project))).replace("\\", "/")
                        except ValueError:
                            dom_rel = str(dom_file)
                except (OSError, RuntimeError, AttributeError, TypeError, ValueError):
                    pass
            return b64, shot_rel, dom_rel
        except OSError:
            return b64, "", ""


# ============================================================================
# 运行期辅助：循环分组 / else 分隔 / 比较算子 / 表格读取
# ============================================================================

def _load_inner_case(path: str) -> TestCase:
    """加载被内嵌引用的用例（新格式 .tc.yaml 或既有 .tc）。"""
    if path.endswith((".tc.yaml", ".tc.yml")):
        return serializer.load(path)
    return load_testcase(path)


@dataclass
class LoopBlock:
    start: Step                 # keyword_loop_start / mobile_loop_start
    body: list = field(default_factory=list)


def _group_loops(steps: list) -> list:
    """把扁平步骤序列里的 loop_start..loop_end 折叠成 LoopBlock（支持嵌套）。"""
    out: list = []
    i = 0
    n = len(steps)
    while i < n:
        node = steps[i]
        if isinstance(node, Step) and node.keyword_id in LOOP_START_IDS:
            depth = 1
            j = i + 1
            while j < n and depth > 0:
                nj = steps[j]
                if isinstance(nj, Step) and nj.keyword_id in LOOP_START_IDS:
                    depth += 1
                elif isinstance(nj, Step) and nj.keyword_id in LOOP_END_IDS:
                    depth -= 1
                    if depth == 0:
                        break
                j += 1
            body = steps[i + 1:j]      # 不含 start / end
            out.append(LoopBlock(start=node, body=body))
            i = j + 1                  # 跳过 end
        elif isinstance(node, Step) and node.keyword_id in LOOP_END_IDS:
            i += 1                     # 落单的 end，忽略
        else:
            out.append(node)
            i += 1
    return out


def _split_else(children: list) -> tuple[list, list]:
    """按 `else` 标记把条件子步骤拆成 (if 体, else 体)。"""
    for idx, c in enumerate(children):
        if isinstance(c, Step) and c.keyword_id in ELSE_IDS:
            return children[:idx], children[idx + 1:]
    return list(children), []


def _eval_expr(expr: str) -> bool:
    """表达式条件（兼容旧样例 a==b 形式）。"""
    for op in ("==", "!=", ">=", "<=", ">", "<"):
        if op in expr:
            a, b = (s.strip() for s in expr.split(op, 1))
            if op == "==":
                return a == b
            if op == "!=":
                return a != b
            try:
                fa, fb = float(a), float(b)
            except ValueError:
                return False
            return {">=": fa >= fb, "<=": fa <= fb, ">": fa > fb, "<": fa < fb}[op]
    return expr.strip().lower() in ("true", "1", "yes")


def _compare(v1: str, v2: str, op: str) -> bool:
    """expResult 算子比较。"""
    if op.startswith("等于"):
        return v1 == v2
    if op == "不等于":
        return v1 != v2
    if op.startswith("模糊匹配"):
        return v2 in v1
    if op == "多值匹配":
        return v1 in [s.strip() for s in v2.split(",")]
    if op in ("大于", "小于", "大于等于", "小于等于"):
        try:
            f1, f2 = float(v1), float(v2)
        except ValueError:
            return False
        return {"大于": f1 > f2, "小于": f1 < f2,
                "大于等于": f1 >= f2, "小于等于": f1 <= f2}[op]
    return v1 == v2


def _parse_datatable(spec: str) -> tuple[str, bool]:
    """解析数据池绑定 `DATATABLE(源,私有)` → (源, is_private)。

    源 = 数据文件路径(.xlsx/.csv/.json)或 'NONE'（不绑定）；私有 = true/false。
    非 DATATABLE(...) 形式一律视作未绑定。以最后一个逗号切分私有标志，容忍源里含逗号。
    """
    s = (spec or "").strip()
    m = re.match(r"^DATATABLE\s*\((.*)\)\s*$", s, re.IGNORECASE | re.DOTALL)
    if not m:
        return "", False
    inner = m.group(1).strip()
    if "," in inner:
        source, priv = inner.rsplit(",", 1)
        return source.strip(), priv.strip().lower() in ("true", "1", "yes", "t")
    return inner, False


def datatable_columns(spec: str, base_dir: str = "") -> list[str]:
    """按数据池绑定 `DATATABLE(源,私有)` 读取数据文件表头列名（供参数 COLUMN 选择器）。

    未绑定/NONE/文件不存在/读取失败一律返回 []。相对路径按 base_dir 解析。
    """
    src, _priv = _parse_datatable(spec)
    if not src or src.upper() == "NONE":
        return []
    path = src
    if not os.path.isabs(path) and base_dir:
        cand = os.path.join(base_dir, path)
        if os.path.exists(cand):
            path = cand
    if not os.path.exists(path):
        return []
    # noinspection PyBroadException
    try:
        if path.lower().endswith(".json"):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return list(dict(data[0]).keys()) if data else []
        rows = _read_table(path)
        return list(rows[0].keys()) if rows else []
    except Exception:
        return []


def _read_table(path: str) -> list[dict]:
    """读取 Excel(.xlsx)/CSV 为行 dict 列表（首行为表头）。"""
    if path.lower().endswith((".xlsx", ".xlsm")):
        # noinspection PyUnresolvedReferences
        from openpyxl import load_workbook  # 延迟：可选 extra，仅 Excel 数据池
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        out = []
        for r in rows[1:]:
            out.append({headers[i]: ("" if v is None else v) for i, v in enumerate(r)
                        if i < len(headers)})
        return out
    # CSV
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]
