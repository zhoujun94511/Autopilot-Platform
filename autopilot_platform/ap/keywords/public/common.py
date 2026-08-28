"""Public / Common 通用工具关键字。

关键字 id 见 keyword_defs 定义（参考 reverse/docs/manifests/Public_CommonKeyword.json，51 条），行为依据
reverse/keyword/sources/CommonKeyword.java。
绝大多数为无外部依赖的纯工具，用 Python 标准库实现；依赖原生/COM/图像比对等
极少数能力降级为 NotImplementedKeyword 但仍注册。

约定：
- out 参数的「值」即要写入的变量名，函数 return {<形参值>: 结果}。
- 末尾统一 **params 兜底。
- ctx 为 ExecutionContext。
"""

from __future__ import annotations

import base64
import csv as _csv
import hashlib
import os
import random
import re
import socket
import time
import urllib.parse
import uuid
from datetime import datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal

from ..registry import keyword, KeywordError, NotImplementedKeyword
from ..context import ExecutionContext
from ...runtime.paths import join_project, to_native


# --------------------------------------------------------------------------- #
# 内部辅助
# --------------------------------------------------------------------------- #
def _to_int(v, default: int = 0) -> int:
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        return default


def _java_fmt_to_py(fmt: str) -> str:
    """把 Java SimpleDateFormat 模式转成 Python strftime 模式（覆盖常用字段）。"""
    if not fmt:
        return "%Y-%m-%d %H:%M:%S"
    # 顺序敏感：先长后短
    table = [
        ("yyyy", "%Y"),
        ("yy", "%y"),
        ("MM", "%m"),
        ("dd", "%d"),
        ("HH", "%H"),
        ("hh", "%I"),
        ("mm", "%M"),
        ("ss", "%S"),
        ("SSS", "%f"),  # 毫秒→微秒(后续截断)
        ("EEE", "%a"),
        ("MMM", "%b"),
        ("zzz", "%Z"),
        ("a", "%p"),
    ]
    out = fmt
    for j, p in table:
        out = out.replace(j, p)
    return out


def _strftime(dt: datetime, fmt: str) -> str:
    py = _java_fmt_to_py(fmt)
    s = dt.strftime(py)
    if "%f" in py:  # 微秒(6位)截成毫秒(3位)
        s = re.sub(r"(\d{3})\d{3}", r"\1", s)
    return s


def _compare_num(d1, d2, exp: str) -> bool:
    """数值比较：exp ∈ 大于/等于/小于/大于等于/小于等于/不等于。"""
    a, b = Decimal(str(d1)), Decimal(str(d2))
    return {
        "大于": a > b,
        "等于": a == b,
        "小于": a < b,
        "大于等于": a >= b,
        "小于等于": a <= b,
        "不等于": a != b,
    }[exp]


def _resolve_proj_path(ctx: ExecutionContext, file_position: str, file_path: str) -> str:
    """工程/本地定位：工程相对 ctx.variables['__project_path__']（若有），本地按原样。"""
    fp = to_native(file_path)
    if file_position == "工程":
        base = ctx.variables.get("__project_path__", "")
        return join_project(base, fp) if base else fp
    return fp


# --------------------------------------------------------------------------- #
# 字符串处理
# --------------------------------------------------------------------------- #
@keyword("common_verify_String", name="字符串校验", category="Public",
         legacy_impl="CommonKeyword:verifyString")
def verify_string(_ctx: ExecutionContext, text="", expect="", matched="true",
                  mode="精确匹配", **_kw):
    """按 mode 校验 text 与 expect 的关系，matched=false 时取反。"""
    text, expect = str(text), str(expect)
    if mode == "精确匹配":
        ok = text == expect
    elif mode == "模糊匹配":
        ok = expect in text
    elif mode == "多值匹配":
        ok = all(v in text for v in expect.split(","))
    elif mode == "正则表达式匹配":
        ok = re.search(expect, text) is not None
    elif mode == "忽略大小写":
        ok = text.lower() == expect.lower()
    elif mode in ("大于", "小于", "大于等于", "小于等于"):
        ok = _compare_num(text, expect, mode)
    else:
        raise KeywordError(f"不支持的匹配模式: {mode}")
    want = str(matched).lower() == "true"
    if ok != want:
        raise KeywordError(f"字符串校验失败：mode={mode} text=[{text}] expect=[{expect}] matched={matched}")


# noinspection PyPep8Naming
@keyword("common_sreplace_Str", name="字符串替换", category="Public",
         out_params=["toStr"], legacy_impl="CommonKeyword:replaceStr")
def replace_str(_ctx: ExecutionContext, fromStr="", oldChar="", newChar="",
                toStr="var_repValue", **_kw):
    """将 fromStr 中的 oldChar 全部替换为 newChar；STR.EMPTY 视作空格。"""
    if oldChar == "STR.EMPTY":
        oldChar = " "
    if newChar == "STR.EMPTY":
        newChar = " "
    return {toStr: str(fromStr).replace(oldChar, newChar)}


# noinspection PyPep8Naming
@keyword("common_trim_str", name="字符串截空", category="Public",
         out_params=["toStr"], legacy_impl="CommonKeyword:trimStr")
def trim_str(_ctx: ExecutionContext, fromStr="", toStr="var_trimValue", **_kw):
    """去除首尾空白。"""
    return {toStr: None if fromStr is None else str(fromStr).strip()}


# noinspection PyPep8Naming
@keyword("common_create_RandomNum", name="字符串随机生成", category="Public",
         out_params=["randomNum"], legacy_impl="CommonKeyword:creatRandomNum")
def create_random_num(_ctx: ExecutionContext, length="20", randomNum="var_randomNum", **_kw):
    """生成指定长度的随机数字字符串。"""
    n = _to_int(length, 20)
    s = "".join(random.choice("0123456789") for _ in range(n))
    return {randomNum: s}


@keyword("common_generate_empty_str", name="生成指定长度空字符串", category="Public",
         out_params=["reference"], legacy_impl="CommonKeyword:generateEmptyStr")
def generate_empty_str(_ctx: ExecutionContext, length="1", reference="", **_kw):
    """生成由空格组成的指定长度字符串。"""
    return {reference: " " * _to_int(length, 1)}


# noinspection PyPep8Naming
@keyword("common_subString_BetweenBeginAndEnd", name="截取从指定开始位至结束位间的字符串",
         category="Public", out_params=["sepValue"],
         legacy_impl="CommonKeyword:subStringBetweenBeginAndEnd")
def substring_between(_ctx: ExecutionContext, string="", beginIndex="", endIndex="",
                      sepValue="var_sepValue", **_kw):
    """str[begin:end]。"""
    return {sepValue: str(string)[_to_int(beginIndex):_to_int(endIndex)]}


# noinspection PyPep8Naming
@keyword("common_subString_ByLength", name="截取从指定位开始及长度的字符串",
         category="Public", out_params=["sepValue"],
         legacy_impl="CommonKeyword:subStringByLength")
def substring_by_length(_ctx: ExecutionContext, string="", beginIndex="", length="",
                        sepValue="var_sepValue", **_kw):
    """str[begin:begin+length]。"""
    b = _to_int(beginIndex)
    return {sepValue: str(string)[b:b + _to_int(length)]}


# noinspection PyPep8Naming
@keyword("common_subString_ByBegin", name="截取从指定位开始至末尾的全部字符",
         category="Public", out_params=["sepValue"],
         legacy_impl="CommonKeyword:subStringByBegin")
def substring_by_begin(_ctx: ExecutionContext, string="", beginIndex="",
                       sepValue="var_sepValue", **_kw):
    """str[begin:]。"""
    return {sepValue: str(string)[_to_int(beginIndex):]}


# noinspection PyPep8Naming
@keyword("common_split_AndGetLength", name="分割字符串并获取长度", category="Public",
         out_params=["length"], legacy_impl="CommonKeyword:splitAndGetLength")
def split_and_get_length(_ctx: ExecutionContext, sourceStr="", splitStr="",
                         length="var_length", **_kw):
    """按 splitStr 分割并返回片段数。"""
    parts = str(sourceStr).split(splitStr) if splitStr else [str(sourceStr)]
    return {length: str(len(parts))}


# noinspection PyPep8Naming
@keyword("common_split_AndGetValue", name="分割字符串并获取列值", category="Public",
         out_params=["value"], legacy_impl="CommonKeyword:splitAndGetValue")
def split_and_get_value(_ctx: ExecutionContext, sourceStr="", splitStr="", col="1",
                        value="var_value", **_kw):
    """按 splitStr 分割，取第 col 列（1 起）。"""
    parts = str(sourceStr).split(splitStr) if splitStr else [str(sourceStr)]
    idx = _to_int(col, 1) - 1
    if idx < 0 or idx >= len(parts):
        raise KeywordError(f"分割取列越界：col={col} 实际片段数={len(parts)}")
    return {value: parts[idx]}


# noinspection PyShadowingBuiltins
@keyword("common_get_str_length", name="获取字符串长度", category="Public",
         out_params=["length"], legacy_impl="CommonKeyword:getStrLength")
def get_str_length(_ctx: ExecutionContext, str="", length="var_length", **_kw):
    """返回字符串长度（None 记 0）。"""
    return {length: "0" if str is None else f"{len(str)}"}


# noinspection PyShadowingBuiltins
@keyword("common_string_case_transform", name="字符串大小写转换", category="Public",
         out_params=["value"], legacy_impl="CommonKeyword:caseTransform")
def case_transform(_ctx: ExecutionContext, str="", type="UP", value="var_value", **_kw):
    """type 含 up→转大写，否则转小写。"""
    t = (type or "up").lower()
    return {value: str.upper() if "up" in t else str.lower()}


# noinspection PyPep8Naming,PyShadowingBuiltins
@keyword("common_compare_str_length", name="校验字符串长度", category="Public",
         legacy_impl="CommonKeyword:compareStrLength")
def compare_str_length(_ctx: ExecutionContext, str="", length="", expResult="等于", **_kw):
    """校验 len(str) 与 length 的关系。"""
    if not _compare_num(len(str), length, expResult):
        raise KeywordError(f"字符串长度校验失败：len={len(str)} 期望{expResult}{length}")


# --------------------------------------------------------------------------- #
# URL / 表单
# --------------------------------------------------------------------------- #
def _url_value(url: str, key: str) -> str:
    """从 url 的 query 中取 key 的值。"""
    q = urllib.parse.urlparse(url).query or url
    params = urllib.parse.parse_qs(q, keep_blank_values=True)
    if key in params:
        return params[key][0]
    return ""


@keyword("common_get_url_element", name="获取URL元素内容", category="Public",
         out_params=["value"], legacy_impl="CommonKeyword:geturlelement")
def get_url_element(_ctx: ExecutionContext, url="", key="", value="", **_kw):
    """提取 url 中参数 key 的值。"""
    return {value: _url_value(url, key)}


# noinspection PyShadowingBuiltins
@keyword("form_compare", name="校验URL元素内容", category="Public",
         legacy_impl="CommonKeyword:formcompare")
def form_compare(_ctx: ExecutionContext, str="", key="", expvalue="", **_kw):
    """校验 str(表单/URL) 中 key 的值等于 expvalue。"""
    actual = _url_value(str, key)
    if actual != expvalue:
        raise KeywordError(f"URL元素校验失败：key={key} 期望[{expvalue}] 实际[{actual}]")


# --------------------------------------------------------------------------- #
# 业务元数据 / 文件
# --------------------------------------------------------------------------- #
# noinspection PyPep8Naming
@keyword("df_define_metadata", name="定义业务元数据", category="Public",
         legacy_impl="CommonKeyword:defineDfMetadata")
def define_metadata(ctx: ExecutionContext, metadataName="", metadataValue="", **_kw):
    """定义业务元数据（暂存入 ctx，仅 TestRunner 端持久化，此处记录）。"""
    store = ctx.variables.setdefault("__df_metadata__", {})
    store[metadataName] = metadataValue
    ctx.log(f"定义业务元数据 {metadataName}={metadataValue}")


# noinspection PyShadowingBuiltins
@keyword("read_file", name="读取文件内容", category="Public",
         out_params=["out"], legacy_impl="CommonKeyword:readFile")
def read_file(ctx: ExecutionContext, path="", format="UTF-8", out="out", **_kw):
    """读取文本文件全部内容（按行拼接，每行追加换行），存入 out 变量。"""
    p = to_native(path)
    base = ctx.variables.get("__project_path__", "")
    full = join_project(base, p) if base else p
    enc = (format or "UTF-8").replace("UTF-8", "utf-8")
    with open(full, "r", encoding=enc) as f:
        content = "".join(line.rstrip("\n") + "\n" for line in f)
    return {out: content}


# --------------------------------------------------------------------------- #
# 时间 / 时间戳
# --------------------------------------------------------------------------- #
@keyword("web_common_sleep", name="等待时间", category="Public",
         legacy_impl="CommonKeyword:sleep")
def sleep(_ctx: ExecutionContext, millis="", **_kw):
    """休眠指定毫秒。"""
    time.sleep(_to_int(millis, 0) / 1000.0)


# noinspection PyShadowingBuiltins
@keyword("common_generate_timestamp", name="生成当前时间戳字符", category="Public",
         out_params=["reference"], legacy_impl="CommonKeyword:generateTimestamp")
def generate_timestamp(_ctx: ExecutionContext, format="yyyy-MM-dd HH:mm:ss",
                       year="0", month="0", day="0", hour="0", minute="0",
                       second="0", reference="", **_kw):
    """以当前时间为基准做年月日时分秒偏移后按 format 格式化。"""
    dt = datetime.now()
    months = _to_int(month)
    # 月份偏移
    if months:
        total = (dt.year * 12 + dt.month - 1) + months
        dt = dt.replace(year=total // 12, month=total % 12 + 1)
    dt = dt + timedelta(days=365 * _to_int(year) + _to_int(day),
                        hours=_to_int(hour), minutes=_to_int(minute),
                        seconds=_to_int(second))
    return {reference: _strftime(dt, format)}


# noinspection PyShadowingBuiltins,PyShadowingNames
@keyword("common_get_timestamp", name="生成指定时间戳字符串", category="Public",
         out_params=["reference"], legacy_impl="CommonKeyword:getTimestamp")
def get_timestamp(_ctx: ExecutionContext, format="yyyy-MM-dd HH:mm:ss",
                  time="2018-04-24 00:00:00", reference="", **_kw):
    """将标准时间字符串转为 Unix 秒时间戳字符串。"""
    dt = datetime.strptime(time, _java_fmt_to_py(format))
    return {reference: str(int(dt.timestamp()))}


# noinspection PyShadowingBuiltins,PyShadowingNames
@keyword("common_get_timestamp_ms", name="生成指定时间戳字符串(支持毫秒)", category="Public",
         out_params=["reference"], legacy_impl="CommonKeyword:getTimestampForMS")
def get_timestamp_ms(_ctx: ExecutionContext, format="yyyy-MM-dd HH:mm:ss:SSS",
                     time="2018-04-24 12:20:05:123", reference="", **_kw):
    """将带毫秒的时间字符串转为 Unix 毫秒时间戳字符串。"""
    dt = datetime.strptime(time, _java_fmt_to_py(format))
    return {reference: str(int(dt.timestamp() * 1000))}


# noinspection PyShadowingBuiltins,PyShadowingNames
@keyword("common_get_timestamp_fromDate", name="从Date格式字符串生成Date数据", category="Public",
         out_params=["reference"], legacy_impl="CommonKeyword:getDateFromDateString")
def get_date_from_string(_ctx: ExecutionContext,
                         format="EEE MMM dd HH:mm:ss zzz yyyy",
                         time="Fri Sep 07 14:21:43 CST 2018", reference="", **_kw):
    """解析 Date 字符串并以标准格式回存（CST 等时区名按本地解析）。"""
    s = re.sub(r"\b[A-Z]{3,4}\b(?=\s+\d{4}$)", "", time).strip()
    fmt = _java_fmt_to_py(format).replace("%Z", "").strip()
    try:
        dt = datetime.strptime(s, fmt)
    except ValueError:
        dt = datetime.strptime(s, "%a %b %d %H:%M:%S %Y")
    return {reference: _strftime(dt, "yyyy-MM-dd HH:mm:ss")}


# noinspection PyPep8Naming,PyShadowingNames
@keyword("common_convert_time_format", name="时间格式转换", category="Public",
         out_params=["outVar"], legacy_impl="CommonKeyword:convertTimeFormat")
def convert_time_format(_ctx: ExecutionContext, format_bef="yyyy-MM-dd HH:mm:ss",
                        time="", format_aft="yyyy-MM-dd HH:mm:ss", outVar="", **_kw):
    """在标准格式、秒时间戳、毫秒时间戳之间互转。"""
    SEC = "时间戳字符串（秒）"
    MS = "时间戳字符串（毫秒）"

    def to_dt_from_stamp(s, ms=False):
        v = int(s)
        return datetime.fromtimestamp(v / 1000.0 if ms else v)

    if format_bef == SEC:
        if not str(time).strip().lstrip("-").isdigit():
            raise KeywordError(f"输入的时间戳秒格式{time}格式不正确")
        if format_aft == SEC:
            res = str(time)
        elif format_aft == MS:
            res = str(time) + "000"
        else:
            res = _strftime(to_dt_from_stamp(time), format_aft)
    elif format_bef == MS:
        if not str(time).strip().lstrip("-").isdigit():
            raise KeywordError(f"输入的时间戳毫秒格式{time}格式不正确")
        if format_aft == SEC:
            res = str(time)[:-3]
        elif format_aft == MS:
            res = str(time)
        else:
            res = _strftime(to_dt_from_stamp(time, ms=True), format_aft)
    else:
        dt = datetime.strptime(time, _java_fmt_to_py(format_bef))
        if format_aft == SEC:
            res = str(int(dt.timestamp()))
        elif format_aft == MS:
            res = str(int(dt.timestamp() * 1000))
        else:
            res = _strftime(dt, format_aft)
    return {outVar: res}


# noinspection PyPep8Naming
@keyword("common_get_current_millis", name="获取当前毫秒时间戳", category="Public",
         out_params=["timeMillis"], legacy_impl="CommonKeyword:getCurrentTimeMillis")
def get_current_millis(_ctx: ExecutionContext, timeMillis="VAR_VALUE", **_kw):
    """返回当前毫秒时间戳。"""
    return {timeMillis: str(int(time.time() * 1000))}


# --------------------------------------------------------------------------- #
# 网络 / 随机生成
# --------------------------------------------------------------------------- #
# noinspection PyPep8Naming
@keyword("common_get_LocalHost_IPAddress", name="获取本机IP地址", category="Public",
         out_params=["localIP"], legacy_impl="CommonKeyword:getLocalHostIPAddress")
def get_local_ip(_ctx: ExecutionContext, localIP="var_localIP", **_kw):
    """获取本机 IP 地址。"""
    try:
        ip = socket.gethostbyname(socket.gethostname())
    except OSError:
        ip = "127.0.0.1"
    return {localIP: ip}


@keyword("common_get_EmailAddress", name="邮箱地址随机生成", category="Public",
         out_params=["email"], legacy_impl="CommonKeyword:getEmailAddress")
def get_email(_ctx: ExecutionContext, internet="@example.com", email="var_email", **_kw):
    """随机生成邮箱地址（前缀随机 + 指定域名）。"""
    prefix = "".join(random.choice("abcdefghijklmnopqrstuvwxyz0123456789") for _ in range(10))
    return {email: prefix + internet}


@keyword("common_get_TelephoneNumber", name="手机号码随机生成", category="Public",
         out_params=["phone"], legacy_impl="CommonKeyword:getTelephoneNumber")
def get_phone(_ctx: ExecutionContext, phone="var_phone", **_kw):
    """随机生成 11 位手机号（1 开头）。"""
    num = "1" + random.choice("3456789") + "".join(random.choice("0123456789") for _ in range(9))
    return {phone: num}


# noinspection PyPep8Naming
@keyword("id_card_creat", name="生成身份证号", category="Public",
         out_params=["idCardNum"], legacy_impl="CommonKeyword:idCardCreat")
def id_card(_ctx: ExecutionContext, idCardNum="var_idCardNum", **_kw):
    """随机生成 18 位身份证号（含校验位）。"""
    area = random.choice(["110101", "310101", "320101", "440101", "510101"])
    year = random.randint(1960, 2005)
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    seq = "%03d" % random.randint(1, 999)
    body = f"{area}{year:04d}{month:02d}{day:02d}{seq}"
    weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    check = "10X98765432"[sum(int(c) * w for c, w in zip(body, weights)) % 11]
    return {idCardNum: body + check}


@keyword("common_get_36UID", name="获取uuid(36位)", category="Public",
         out_params=["uuid_36"], legacy_impl="CommonKeyword:getUUID")
def get_uuid(_ctx: ExecutionContext, uuid_36="", **_kw):
    """生成 36 位标准 UUID 字符串。"""
    return {uuid_36: str(uuid.uuid4())}


# --------------------------------------------------------------------------- #
# 数值计算 / 比较
# --------------------------------------------------------------------------- #
# noinspection PyPep8Naming
@keyword("common_data_calc", name="数值计算", category="Public",
         out_params=["targetData"], legacy_impl="CommonKeyword:dataCalc")
def data_calc(_ctx: ExecutionContext, data1="", data2="", data3="0",
              Decimal_range="除不尽场景", operator="加", targetData="var_calcValue", **_kw):
    """加减乘除取余；data3 为保留小数位，除法 Decimal_range 控制是否截断。"""
    a, b = Decimal(str(data1)), Decimal(str(data2))
    scale = _to_int(data3, 0)
    if operator == "加":
        r = a + b
    elif operator == "减":
        r = a - b
    elif operator == "乘":
        r = a * b
    elif operator == "除":
        if b == 0:
            raise KeywordError("除数不能为0")
        r = a / b
    elif operator == "取余":
        r = a % b
    else:
        raise KeywordError(f"不支持的运算符: {operator}")
    if operator == "除" and Decimal_range == "除法计算":
        # 按保留位数四舍五入
        q = Decimal(1).scaleb(-scale)
        r = r.quantize(q, rounding=ROUND_HALF_UP)
    elif scale > 0:
        q = Decimal(1).scaleb(-scale)
        r = r.quantize(q, rounding=ROUND_HALF_UP)
    return {targetData: str(r.normalize() if scale == 0 else r)}


# noinspection PyPep8Naming
@keyword("roundValue", name="四舍五入", category="Public",
         out_params=["targetData"], legacy_impl="CommonKeyword:roundValue")
def round_value(_ctx: ExecutionContext, value="", length="0", targetData="var_calcValue", **_kw):
    """对 value 保留 length 位小数（四舍五入）。"""
    scale = _to_int(length, 0)
    q = Decimal(1).scaleb(-scale)
    r = Decimal(str(value)).quantize(q, rounding=ROUND_HALF_UP)
    return {targetData: str(r)}


# noinspection PyPep8Naming
@keyword("common_data_compare", name="数值比较", category="Public",
         legacy_impl="CommonKeyword:compareToData")
def data_compare(_ctx: ExecutionContext, data1="", data2="", expResult="大于", **_kw):
    """校验 data1 与 data2 的数值关系。"""
    if not _compare_num(data1, data2, expResult):
        raise KeywordError(f"数值比较失败：{data1} 期望{expResult} {data2}")


# noinspection PyPep8Naming
@keyword("common_data_between", name="数值范围校验", category="Public",
         legacy_impl="CommonKeyword:dataBetween")
def data_between(_ctx: ExecutionContext, data="", expResult1="大于", boundary1="0",
                 expResult2="小于", boundary2="0", **_kw):
    """校验 data 同时满足与 boundary1、boundary2 的两个关系。"""
    if not _compare_num(data, boundary1, expResult1):
        raise KeywordError(f"数值范围校验失败：{data} 期望{expResult1} {boundary1}")
    if not _compare_num(data, boundary2, expResult2):
        raise KeywordError(f"数值范围校验失败：{data} 期望{expResult2} {boundary2}")


# --------------------------------------------------------------------------- #
# 变量 / 日志
# --------------------------------------------------------------------------- #
# noinspection PyPep8Naming
@keyword("setVariable", name="自定义变量", category="Public",
         out_params=["varName"], legacy_impl="CommonKeyword:setVariable")
def set_variable(_ctx: ExecutionContext, varName="", varValue="", **_kw):
    """把 varValue 存入名为 varName 的变量。"""
    return {varName: varValue}


# noinspection PyPep8Naming
@keyword("logPrint", name="日志打印", category="Public",
         legacy_impl="CommonKeyword:logPrint")
def log_print(ctx: ExecutionContext, logLevel="info", message="", **_kw):
    """按级别打印日志。"""
    ctx.log(f"[{logLevel}] {message}")


# --------------------------------------------------------------------------- #
# 编码 / 加密
# --------------------------------------------------------------------------- #
@keyword("getMd5", name="MD5加密", category="Public",
         out_params=["result"], legacy_impl="CommonKeyword:getMd5")
def get_md5(_ctx: ExecutionContext, string="", result="", **_kw):
    """计算字符串 MD5（小写十六进制）。"""
    return {result: hashlib.md5(str(string).encode("utf-8")).hexdigest()}


@keyword("base64加密", name="base64加密", category="Public",
         out_params=["result"], legacy_impl="CommonKeyword:base64Encode")
def base64_encode(_ctx: ExecutionContext, string="", charset="GB18030", result="", **_kw):
    """按指定字符集 base64 编码。"""
    cs = (charset or "GB18030")
    return {result: base64.b64encode(str(string).encode(cs)).decode("ascii")}


@keyword("base64解密", name="base64解密", category="Public",
         out_params=["result"], legacy_impl="CommonKeyword:base64Decode")
def base64_decode(_ctx: ExecutionContext, string="", charset="GB18030", result="", **_kw):
    """按指定字符集 base64 解码。"""
    cs = (charset or "GB18030")
    return {result: base64.b64decode(str(string).encode("ascii")).decode(cs)}


# noinspection PyShadowingBuiltins
@keyword("URLcode", name="URL编码/解码", category="Public",
         out_params=["result"], legacy_impl="CommonKeyword:urlCode")
def url_code(_ctx: ExecutionContext, string="", type="解码", charset="UTF-8",
             result="var_CodedUrl", **_kw):
    """type=编码→URL 编码，否则 URL 解码。"""
    cs = (charset or "UTF-8")
    if type == "编码":
        val = urllib.parse.quote_plus(str(string), encoding=cs)
    else:
        val = urllib.parse.unquote_plus(str(string), encoding=cs)
    return {result: val}


# --------------------------------------------------------------------------- #
# 多结果校验（与）
# --------------------------------------------------------------------------- #
def _multi_and(ctx: ExecutionContext, multi_param: str) -> bool:
    """对 && 分隔的多个变量名取值做布尔与。"""
    result = True
    for name in str(multi_param).split("&&"):
        name = name.strip()
        if not name:
            continue
        v = ctx.get_var(name, None)
        if v is not None:
            result = (str(v).lower() == "true") and result
        if not result:
            break
    return result


# noinspection PyPep8Naming
@keyword("exec_control_multiple", name="多结果校验(与)", category="Public",
         legacy_impl="CommonKeyword:multiVariableVerify")
def multi_verify(ctx: ExecutionContext, multiParam="", expResult="true", **_kw):
    """对多变量做与运算并与期望结果比较。"""
    result = _multi_and(ctx, multiParam)
    if str(result).lower() != str(expResult).lower():
        raise KeywordError(f"多变量校验失败，期望{expResult}，实际{result}")


# noinspection PyPep8Naming
@keyword("exec_set_control_multiple", name="获取与结果", category="Public",
         out_params=["outVar"], legacy_impl="CommonKeyword:setMultiVariableVerifyStatus")
def set_multi_verify(ctx: ExecutionContext, multiParam="", outVar="", **_kw):
    """对多变量做与运算并把布尔结果存入 outVar。"""
    return {outVar: str(_multi_and(ctx, multiParam)).lower()}


# --------------------------------------------------------------------------- #
# 终端类型
# --------------------------------------------------------------------------- #
# noinspection PyPep8Naming
@keyword("common_set_terminal_Type", name="测试终端类型手工设置", category="Public",
         legacy_impl="CommonKeyword:setTerminalType")
def set_terminal_type(ctx: ExecutionContext, terminalType="PC", **_kw):
    """设置测试终端类型（PC/Mobile），存入上下文。"""
    ctx.set_var("__terminal_type__", terminalType)


@keyword("common_reset_terminal_Type", name="取消测试终端类型设置", category="Public",
         legacy_impl="CommonKeyword:resetTerminalType")
def reset_terminal_type(ctx: ExecutionContext, **_kw):
    """清除测试终端类型设置。"""
    ctx.variables.pop("__terminal_type__", None)


# --------------------------------------------------------------------------- #
# 文件 / 校验
# --------------------------------------------------------------------------- #
# noinspection PyPep8Naming
@keyword("common_verify_file_existed", name="校验文件是否存在", category="Public",
         legacy_impl="CommonKeyword:verifyFileExisted")
def verify_file_existed(ctx: ExecutionContext, filePosition="本地", filePath="",
                        isExisted="true", **_kw):
    """校验文件是否存在与期望一致。"""
    full = _resolve_proj_path(ctx, filePosition, filePath)
    exists = os.path.isfile(full)
    want = str(isExisted).lower() == "true"
    if exists != want:
        raise KeywordError(f"文件存在性校验失败：{full} 实际存在={exists} 期望={want}")


# noinspection PyPep8Naming
@keyword("common_CSVFile_create", name="生成并保存数据至CSV文件", category="Public",
         legacy_impl="CommonKeyword:createCSVToSaveData")
def create_csv(ctx: ExecutionContext, FilePosition="工程", filePath="",
               tableHead="", tableData="", **_kw):
    """生成 CSV 文件：tableHead 为表头，tableData 多行(以;分隔行,以,分隔列)。"""

    full = _resolve_proj_path(ctx, FilePosition, filePath)
    d = os.path.dirname(full)
    if d:
        os.makedirs(d, exist_ok=True)
    with open(full, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        if tableHead:
            w.writerow(tableHead.split(","))
        for row in str(tableData).split(";"):
            if row != "":
                w.writerow(row.split(","))


# --------------------------------------------------------------------------- #
# Excel（openpyxl，1 起索引）
# --------------------------------------------------------------------------- #
def _excel_load(full):
    try:
        # noinspection PyUnresolvedReferences
        import openpyxl
    except ImportError:
        raise NotImplementedKeyword("缺少 openpyxl 依赖，无法操作 Excel")
    return openpyxl.load_workbook(full)


# noinspection PyPep8Naming
@keyword("common_excel_editCellValue", name="修改excel单元格值", category="Public",
         legacy_impl="CommonKeyword:editExcelCellValue")
def excel_edit(ctx: ExecutionContext, filePosition="工程", filePath="", sheetIndex="1",
               rowIndex="1", colIndex="1", value="", **_kw):
    """修改 excel 指定单元格的值并保存（索引 1 起）。"""
    full = _resolve_proj_path(ctx, filePosition, filePath)
    wb = _excel_load(full)
    ws = wb.worksheets[_to_int(sheetIndex, 1) - 1]
    ws.cell(row=_to_int(rowIndex, 1), column=_to_int(colIndex, 1), value=value)
    wb.save(full)


# noinspection PyPep8Naming
@keyword("common_excel_writeCellValue", name="excel单元格写入值", category="Public",
         legacy_impl="CommonKeyword:writeExcelCellValue")
def excel_write(ctx: ExecutionContext, filePosition="工程", filePath="", sheetIndex="1",
                rowIndex="1", colIndex="1", value="", **_kw):
    """向 excel 单元格写入值并保存（索引 1 起）。"""
    full = _resolve_proj_path(ctx, filePosition, filePath)
    wb = _excel_load(full)
    ws = wb.worksheets[_to_int(sheetIndex, 1) - 1]
    ws.cell(row=_to_int(rowIndex, 1), column=_to_int(colIndex, 1), value=value)
    wb.save(full)


# noinspection PyPep8Naming
@keyword("common_excel_readCellValue", name="读取excel单元格值", category="Public",
         out_params=["value"], legacy_impl="CommonKeyword:readExcelCellValue")
def excel_read(ctx: ExecutionContext, filePosition="工程", filePath="", sheetIndex="1",
               rowIndex="1", colIndex="1", value="", **_kw):
    """读取 excel 指定单元格的值（索引 1 起）。"""
    full = _resolve_proj_path(ctx, filePosition, filePath)
    wb = _excel_load(full)
    ws = wb.worksheets[_to_int(sheetIndex, 1) - 1]
    v = ws.cell(row=_to_int(rowIndex, 1), column=_to_int(colIndex, 1)).value
    return {value: "" if v is None else str(v)}


# noinspection PyPep8Naming
@keyword("common_excel_delRow", name="删除excel中某行", category="Public",
         legacy_impl="CommonKeyword:delExcelRow")
def excel_del_row(ctx: ExecutionContext, filePosition="工程", filePath="", sheetIndex="1",
                  rowIndex="1", delType="删除行内容且下行上移", **_kw):
    """删除 excel 某行：上移整行 或 仅清空该行内容。"""
    full = _resolve_proj_path(ctx, filePosition, filePath)
    wb = _excel_load(full)
    ws = wb.worksheets[_to_int(sheetIndex, 1) - 1]
    r = _to_int(rowIndex, 1)
    if delType == "删除行内容且下行上移":
        ws.delete_rows(r, 1)
    else:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c, value=None)
    wb.save(full)


# --------------------------------------------------------------------------- #
# 降级项（依赖原生/图像/浏览器原址请求）
# --------------------------------------------------------------------------- #
# noinspection PyPep8Naming
@keyword("compare_ImageXY", name="图片对比", category="Public",
         legacy_impl="CommonKeyword:compareImageXY")
def compare_image(ctx: ExecutionContext, imgPath1="", imgPath2="", imgStartX="0",
                  imgStartY="0", imgEndX="0", imgEndY="0", expectSimilarity="100", **_kw):
    """两张图片(可选区域)相似度对比，用 opencv。相似度低于期望阈值则校验失败。

    imgStartX/Y/EndX/Y 指定对比区域(均为0时对比整图)；expectSimilarity 期望相似度百分比[0~100]。
    """
    # noinspection PyUnresolvedReferences,PyPackageRequirements
    import cv2

    a = cv2.imread(imgPath1)
    b = cv2.imread(imgPath2)
    if a is None or b is None:
        raise KeywordError(f"图片读取失败: {imgPath1} / {imgPath2}")

    def crop(img):
        x1, y1, x2, y2 = int(imgStartX), int(imgStartY), int(imgEndX), int(imgEndY)
        return img[y1:y2, x1:x2] if (x2 > x1 and y2 > y1) else img

    import numpy as np

    a, b = crop(a), crop(b)
    h = min(a.shape[0], b.shape[0])
    w = min(a.shape[1], b.shape[1])
    a = cv2.resize(a, (w, h)).astype("float32")
    b = cv2.resize(b, (w, h)).astype("float32")
    # 平均绝对差归一化为相似度（对纯色/低方差图也稳定）
    similarity = (1.0 - float(np.abs(a - b).mean()) / 255.0) * 100
    ctx.log(f"图片相似度: {similarity:.2f}% (期望≥{expectSimilarity}%)")
    if similarity < float(expectSimilarity):
        raise KeywordError(f"图片对比失败：相似度 {similarity:.2f}% < 期望 {expectSimilarity}%")


@keyword("ommon_pic_checkPicIsExisted", name="获取页面图片原址请求响应码", category="Public",
         out_params=["errorURL"], legacy_impl="CommonKeyword:checkPicIsExisted")
def check_pic_existed(ctx: ExecutionContext, **_kw):
    """检查当前页面所有 <img> 是否加载成功，未成功的原址写入输出变量 errorURL。

    以浏览器渲染结果为准：img.complete 且 naturalWidth>0 视为加载成功；否则判为失败
    （断链/403/404/超时等）——比逐个 HTTP 请求更贴近“用户实际看到的图有没有裂”。
    需先执行「浏览器打开」并加载页面（无会话时报错，非降级）。
    """
    from ..web.driver import get_manager as _web_manager
    drv = _web_manager(ctx).driver()
    js = (
        "return Array.from(document.images)"
        ".filter(function(im){return im.src && !(im.complete && im.naturalWidth>0);})"
        ".map(function(im){return im.currentSrc||im.src;});"
    )
    broken = drv.execute_script(js) or []
    error_url = ";".join(dict.fromkeys(broken))   # 去重保序
    if broken:
        ctx.log(f"发现 {len(broken)} 张图片加载失败：{error_url}")
    else:
        ctx.log("页面所有图片均加载成功")
    return {"errorURL": error_url}
